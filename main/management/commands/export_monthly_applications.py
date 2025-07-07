# your_app/management/commands/export_monthly_applications.py

import os
from datetime import date, datetime
from django.core.management.base import BaseCommand
from django.conf import settings
from django.utils import timezone

import openpyxl
from openpyxl.utils import get_column_letter

from main.models import Enrollment


class Command(BaseCommand):
    help = "Export enrollments applied on the first of the month, per center, into Excel files"

    def handle(self, *args, **options):
        today = timezone.localdate()
        first = today.replace(day=1)

        # barcha 1-kun qabul qilingan enrolmentlar
        qs = Enrollment.objects.select_related(
            "user",
            "course__branch__edu_center",
            "course__branch",
        ).filter(applied_at__date=first)

        if not qs.exists():
            self.stdout.write(f"No applications on {first}")
            return

        # guruhlash: edu_center -> list of enrollments
        centers = {}
        for e in qs:
            center = e.course.branch.edu_center
            centers.setdefault(center, []).append(e)

        export_dir = os.path.join(settings.MEDIA_ROOT, "exports")
        os.makedirs(export_dir, exist_ok=True)

        for center, enrolls in centers.items():
            wb = openpyxl.Workbook()
            # headerlar
            headers = [
                "full_name",
                "phone_number",
                "course_name",
                "branch_name",
                "applied_at",
                "course_price",
                "charge_percent",
                "charge",
            ]

            # — “All” varaqqa yozamiz
            ws_all = wb.active
            ws_all.title = "All"
            ws_all.append(headers)
            total_charge_all = 0

            for e in enrolls:
                price = float(e.course.price)
                charge = round(price * 0.03, 2)
                total_charge_all += charge

                ws_all.append([
                    e.user.full_name,
                    e.user.phone_number,
                    e.course.name,
                    e.course.branch.name if e.course.branch else "",
                    timezone.localtime(e.applied_at).isoformat(),
                    price,
                    3,
                    charge,
                ])

            # Total satri
            ws_all.append([""] * (len(headers) - 2) + ["Total", total_charge_all])
            # ustunlarni kengaytirish
            for i in range(1, len(headers) + 1):
                ws_all.column_dimensions[get_column_letter(i)].auto_size = True

            # keyin har filial uchun alohida varaqlar
            branches = {}
            for e in enrolls:
                branch = e.course.branch
                branches.setdefault(branch, []).append(e)

            for branch, blist in branches.items():
                ws = wb.create_sheet(title=branch.name[:31])  # sheet name limit 31
                ws.append(headers)
                total_charge = 0
                for e in blist:
                    price = float(e.course.price)
                    charge = round(price * 0.03, 2)
                    total_charge += charge
                    ws.append([
                        e.user.full_name,
                        e.user.phone_number,
                        e.course.name,
                        branch.name,
                        timezone.localtime(e.applied_at).isoformat(),
                        price,
                        3,
                        charge,
                    ])
                ws.append([""] * (len(headers) - 2) + ["Total", total_charge])
                for i in range(1, len(headers) + 1):
                    ws.column_dimensions[get_column_letter(i)].auto_size = True

            fname = f"{center.id}-{center.name.replace(' ','_')}-{first.isoformat()}-applications.xlsx"
            path = os.path.join(export_dir, fname)
            wb.save(path)
            self.stdout.write(self.style.SUCCESS(
                f"Saved center “{center.name}” to {path} ({len(enrolls)} rows)"
            ))
